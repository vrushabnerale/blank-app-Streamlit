import streamlit as st
import pandas as pd
import re
import tempfile
import os
from openpyxl import load_workbook
import seaborn as sns
import matplotlib.pyplot as plt
import plotly.express as px


# Function to extract the 5/6-digit code from Prozessname
def extract_code(prozess_name):
    pattern = r'\b(HDD|KV|OBW)[-\s]\d{2,3}-\d{2}\b'
    match = re.search(pattern, str(prozess_name))
    return match.group(0) if match else None

# Streamlit App
st.title("AMP Top 10 Analysis")

st.header("Lean Export Filtering")

# File Upload
st.markdown("Upload the **Lean Export** Excel files")
uploaded_files = st.file_uploader("Lean Export Excel files", type=['xlsx'], accept_multiple_files=True)


if uploaded_files:
    # Load files and process them
    st.write("Uploaded Files:")
    for file in uploaded_files:
        st.write(f"- {file.name}")

    # User-defined cutoff for KW Start
    kw_cutoff = st.number_input("KW Cutoff (e.g., 27)", min_value=1, max_value=52, value=27, step=1)

    # Add an input for specific KW Start
    user_kw_start = st.number_input("Enter a specific KW Start value to filter", min_value=1, max_value=52, value=1, step=1)

    # Process the files
    filtered_dataframes = []

    for file in uploaded_files:
        # Create a temporary file to store the uploaded data
        with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
            temp_file.write(file.read())  # Write the uploaded file data to the temp file
            temp_file.seek(0)  # Go back to the start of the file
            
            # Load the Excel file from the temporary file
            df = pd.read_excel(temp_file.name)
        
        # Clean up the temporary file
        os.unlink(temp_file.name)

        # Ensure Startdatum is in datetime format
        df['Startdatum'] = pd.to_datetime(df['Startdatum'], errors='coerce')

        # Filter rows where Startdatum is in the year 2025
        filtered_df = df[df['Startdatum'].dt.year == 2025].copy()

        # Filter for KW Start
        KW_cutoff_df = filtered_df[(filtered_df['KW Start'] <= kw_cutoff) & 
                                   (filtered_df['KW Start'] >= user_kw_start)].copy()

        # Filter Gewerke for specific keywords
        keywords = ['HDD', 'OBW', 'offene Bauweise', 'Kurzvortrieb', 'Mikrotunnel', 'MT']
        pattern = '|'.join(keywords)
        gewerk_filtered_df = KW_cutoff_df[KW_cutoff_df['Gewerk'].str.contains(pattern, case=False, na=False)].copy()

        # Filter Prozessname for specific includes and excludes
        includes = ['Fertigstellung', 'OBW', 'HDD']
        excludes = [
            'Vorarbeit', 'Zuwegung', 'PE-Rohre Schweißen', 'PE-Schweißen',
            'Anbindung', 'Oberboden auftragen', 'Oberbodenauftrag',
            'obw baustraße', 'Deichkreuzung', 'Teil'
        ]

        include_pattern = '|'.join(includes)
        exclude_pattern = '|'.join(excludes)

        prozessname_filtered_df = gewerk_filtered_df[
            gewerk_filtered_df['Prozessname'].str.contains(include_pattern, case=False, na=False) &
            ~gewerk_filtered_df['Prozessname'].str.contains(exclude_pattern, case=False, na=False)
        ].copy()

        # Add columns
        prozessname_filtered_df['NDS/NRW'] = file.name.split('.')[0]
        prozessname_filtered_df['Bauweise\nBereichs-\nerkennung'] = prozessname_filtered_df['Prozessname'].apply(extract_code)

        # Keep specific columns
        columns_to_keep = ['Id', 'Prozessname', 'Startdatum', 'Enddatum', 'Status', 'Dauer', 'Gewerk', 'KW Start', 'KW Ende', 'NDS/NRW', 'Bauweise\nBereichs-\nerkennung']
        prozessname_filtered_df = prozessname_filtered_df[columns_to_keep]

        filtered_dataframes.append(prozessname_filtered_df)

    # Combine all filtered DataFrames
    combined_df = pd.concat(filtered_dataframes, ignore_index=True)

    # Display preview of the combined data
    st.subheader("Preview of Processed Data")
    st.dataframe(combined_df)

    # Analysis 1: Count of Entries by NDS/NRW
    st.subheader("Count of Entries by NDS/NRW")
    counts = combined_df['NDS/NRW'].value_counts()
    st.table(counts)
    st.bar_chart(counts)

    # Save the combined DataFrame to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        combined_df.to_excel(temp_file.name, index=False)
        temp_file_path = temp_file.name

    # Option to download the processed data
    with open(temp_file_path, 'rb') as f:
        st.download_button(
            label="Download Processed Data",
            data=f.read(),
            file_name="combined_filtered_data.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
st.header("Kreuzungspartner Filtering")

st.markdown("Upload the **A-Nord mit Parallelführung_Kreuzungen und RNTK** Excel File")
file_part2 = st.file_uploader("A-Nord mit Parallelführung_Kreuzungen und RNTK Excel File", type=['xlsx'])


if file_part2:
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        temp_file.write(file_part2.read())
        temp_file.seek(0)
        file_path = temp_file.name

    sheet_name = st.text_input("Enter the sheet name (default: 'RNTK A-NORD || BD4')", "RNTK A-NORD || BD4")
    column_to_check = st.text_input("Enter the column name to check for strikethrough:", 
                                    "Baufrei aus Kreuzungssicht (immer unter Einhaltung der Auflagen/Techn. Bestimmungen des Kreuzungspartners)")

    # Load workbook and sheet
    workbook = load_workbook(file_path, data_only=True)
    sheet = workbook[sheet_name]
    strikethrough_rows = []

    header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    column_index = header.index(column_to_check) + 1 if column_to_check in header else None

    if column_index:
        for row in sheet.iter_rows(min_row=2, values_only=False):
            cell = row[column_index - 1]
            if cell.font.strike:
                strikethrough_rows.append(cell.row)

    df_part2 = pd.read_excel(file_path, sheet_name=sheet_name)
    os.unlink(file_path)

    nein_filtered_df = df_part2[df_part2[column_to_check].str.contains('nein', case=False, na=False)]
    filtered_df_part2 = nein_filtered_df[~nein_filtered_df.index.isin([r - 2 for r in strikethrough_rows])]

    columns_to_keep = ['ID', 'PFA', 'Kreuzungspartner', 'Bauweise\nBereichs-\nerkennung', 'Kreuzungsobjekt', column_to_check]
    filtered_df_part2 = filtered_df_part2.reindex(columns=columns_to_keep, fill_value='N/A')

    summary_df = (filtered_df_part2.groupby('Kreuzungspartner')['Bauweise\nBereichs-\nerkennung']
                  .nunique()
                  .reset_index()
                  .rename(columns={'Bauweise\nBereichs-\nerkennung': 'Unique Bauweise Count'})
                  .sort_values(by='Unique Bauweise Count', ascending=True))

    st.subheader("Filtered Data (Part 2)")
    st.dataframe(filtered_df_part2)

    st.subheader("Summary Data (Part 2)")
    st.dataframe(summary_df)

    # Ensure the 'summary_df' is displayed
    st.subheader("Bar Chart: Unique Bauweise Count by Kreuzungspartner")

    # Create a Plotly bar chart
    fig = px.bar(
        summary_df,
        x="Unique Bauweise Count",
        y="Kreuzungspartner",
        orientation="h",
        title="Unique Bauweise Count by Kreuzungspartner",
        labels={"Unique Bauweise Count": "Count", "Kreuzungspartner": "Kreuzungspartner"},
        width=800,  # Set the width of the chart
        height=1000,  # Set the height of the chart
    )

    # Display the chart in Streamlit
    st.plotly_chart(fig, use_container_width=True)

    # Save filtered data to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_filtered_file:
        filtered_df_part2.to_excel(temp_filtered_file.name, index=False)
        temp_filtered_file_path = temp_filtered_file.name

    # Save summary data to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_summary_file:
        summary_df.to_excel(temp_summary_file.name, index=False)
        temp_summary_file_path = temp_summary_file.name

    # Download buttons for the temporary files
    with open(temp_filtered_file_path, 'rb') as f:
        st.download_button("Download Filtered Data (Part 2)", 
                        f.read(), file_name="Kreuzungspartners-KW-xx.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    with open(temp_summary_file_path, 'rb') as f:
        st.download_button("Download Summary Data (Part 2)", 
                        f.read(), file_name="Total-kreuzungspartners-KW-xx.xlsx", 
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        
st.header("Final Part: Merging and Generating the Final Output")

# Ensure Part 1 and Part 2 data exist
if 'combined_df' in locals() and 'filtered_df_part2' in locals():
    # Perform the merge operation
    st.write("Merging combined data with Kreuzungspartner data...")
    merged_data = pd.merge(
        combined_df[['Bauweise\nBereichs-\nerkennung', 'KW Start', 'NDS/NRW']],  # Relevant columns from Part 1
        filtered_df_part2[['ID', 'Bauweise\nBereichs-\nerkennung', 'Kreuzungspartner', 'Kreuzungsobjekt']],  # Relevant columns from Part 2
        on='Bauweise\nBereichs-\nerkennung',
        how='left'
    )

    # Filter rows with non-empty 'Kreuzungspartner'
    filtered_data = merged_data[merged_data['Kreuzungspartner'].notnull()]

    # Select only the necessary columns
    columns_to_keep = [
        'Bauweise\nBereichs-\nerkennung',
        'KW Start',
        'ID',
        'Kreuzungsobjekt',
        'Kreuzungspartner',
        'NDS/NRW'
    ]
    filtered_data = filtered_data[columns_to_keep]

    # Add filter for Kreuzungspartner
    st.subheader("Filter Data by Kreuzungspartner")
    unique_kreuzungspartner = filtered_data['Kreuzungspartner'].unique()
    selected_kreuzungspartner = st.multiselect(
        "Select Kreuzungspartner to filter:", 
        options=unique_kreuzungspartner, 
        default=unique_kreuzungspartner
    )
    
    # Apply the filter
    filtered_data = filtered_data[filtered_data['Kreuzungspartner'].isin(selected_kreuzungspartner)]

    # Sort by KW Start in ascending order
    filtered_data = filtered_data.sort_values(by='KW Start', ascending=True)

    # Display the filtered merged data
    st.subheader("Preview of the Filtered Merged Data")
    st.dataframe(filtered_data)

    # Save the final output to a temporary file
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as temp_file:
        filtered_data.to_excel(temp_file.name, index=False)
        temp_file_path = temp_file.name

    # Provide a download button for the final output
    with open(temp_file_path, 'rb') as f:
        st.download_button(
            label="Download Final Output (Merged Data)",
            data=f.read(),
            file_name="Output-KW-xx.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    # Cleanup the temporary file
    os.unlink(temp_file_path)

else:
    st.warning("Please process data in Part 1 and Part 2 before generating the final output.")

